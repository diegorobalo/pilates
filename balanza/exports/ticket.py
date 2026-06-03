"""
exports/ticket.py — Ticket de pesaje (Ingreso y Egreso)
Diseño: logo + encabezado empresa | secciones con header azul | pesos destacados
"""
import os
from datetime import datetime


def _fmt(val):
    try:
        return f"{float(val):,.1f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "0,0"


def _logo_path():
    """Ruta al logo de Argensun (junto al ejecutable)."""
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, "argensun_logo.png")


def generar_ticket(pesaje: dict, empresa: dict, ruta: str,
                   certificado: dict = None,
                   tipo: str = "egreso") -> str:
    """
    Genera un ticket PDF.
    tipo='egreso'  → ticket completo (ingreso + egreso + pesos)
    tipo='ingreso' → comprobante de ingreso (solo datos de entrada + tara)
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, Image as RLImage)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    os.makedirs(os.path.dirname(os.path.abspath(ruta)), exist_ok=True)

    # ── Colores ──────────────────────────────────────────────────────────
    AZ      = colors.HexColor("#1A5FA8")
    AZ_OSC  = colors.HexColor("#0D3D73")
    AZ_BG   = colors.HexColor("#EBF3FF")
    VERDE   = colors.HexColor("#065F46")
    VERDE_BG= colors.HexColor("#ECFDF5")
    GRIS    = colors.HexColor("#6B7280")
    GRIS_L  = colors.HexColor("#E5E7EB")
    GRIS_F  = colors.HexColor("#F3F4F6")
    NEGRO   = colors.black
    BLANCO  = colors.white
    AM      = colors.HexColor("#F59E0B")

    doc = SimpleDocTemplate(
        ruta, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.2*cm,  bottomMargin=1.5*cm,
    )
    W = A4[0] - 3*cm   # ancho disponible

    def s(size=9, bold=False, align=TA_LEFT, color=NEGRO):
        return ParagraphStyle(
            "_", fontName="Helvetica-Bold" if bold else "Helvetica",
            fontSize=size, textColor=color,
            alignment=align, leading=size * 1.45, spaceAfter=0,
        )

    e = []

    # ════════════════════════════════════════════════════════════════════
    # ENCABEZADO: logo | datos empresa | nro ticket
    # ════════════════════════════════════════════════════════════════════
    nom  = empresa.get("nombre",     "")
    cuit = empresa.get("cuit",       "")
    coda = empresa.get("cod_aduana", "")
    clot = empresa.get("cod_lot",    "")
    dire = empresa.get("direccion",  "")   # ← CORREGIDO (era tel)
    tel  = empresa.get("telefono",   "")   # ← CORREGIDO (era dire)

    # Alto fijo del encabezado para alinear las 3 columnas
    HEADER_H = 2.2*cm

    # Columna 1 — Logo con borde dorado
    logo_path = _logo_path()
    if os.path.exists(logo_path):
        logo_img  = RLImage(logo_path, width=4.8*cm, height=1.88*cm)
        logo_cell = Table([[logo_img]], colWidths=[W * 0.33],
                          rowHeights=[HEADER_H])
        logo_cell.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,-1), BLANCO),
            ("BOX",           (0,0),(-1,-1), 2,  AM),
            ("TOPPADDING",    (0,0),(-1,-1), 4),
            ("BOTTOMPADDING", (0,0),(-1,-1), 4),
            ("LEFTPADDING",   (0,0),(-1,-1), 8),
            ("RIGHTPADDING",  (0,0),(-1,-1), 8),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ]))
    else:
        logo_cell = Table([
            [Paragraph(f"<b>{nom}</b>", s(14, bold=True, color=BLANCO))],
        ], colWidths=[W * 0.33], rowHeights=[HEADER_H])
        logo_cell.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,-1), AZ_OSC),
            ("TOPPADDING",    (0,0),(-1,-1), 15),
            ("BOTTOMPADDING", (0,0),(-1,-1), 15),
            ("LEFTPADDING",   (0,0),(-1,-1), 10),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ]))

    # Columna 2 — Datos empresa sobre fondo azul oscuro
    datos_cell = Table([
        [Paragraph(f"<b>{nom}</b>", s(11, bold=True, color=BLANCO))],
        [Paragraph(f"CUIT: {cuit}", s(8, color=colors.HexColor("#B3D4F0")))],
        [Paragraph(f"Dir: {dire}", s(8, color=colors.HexColor("#B3D4F0")))],
        [Paragraph(f"Tel: {tel}  ·  Cód.Aduana: {coda}  ·  LOT: {clot}",
                   s(8, color=colors.HexColor("#B3D4F0")))],
    ], colWidths=[W * 0.47])
    datos_cell.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), AZ_OSC),
        ("TOPPADDING",    (0,0),(-1,-1), 2),
        ("BOTTOMPADDING", (0,0),(-1,-1), 2),
        ("LEFTPADDING",   (0,0),(-1,-1), 10),
        ("RIGHTPADDING",  (0,0),(-1,-1), 6),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
    ]))

    # Columna 3 — Tipo + Nro ticket en recuadro azul
    tipo_txt = "INGRESO" if tipo == "ingreso" else "EGRESO"
    nro_cell = Table([
        [Paragraph(tipo_txt,
                   s(9, bold=True, align=TA_CENTER,
                     color=colors.HexColor("#B3D4F0")))],
        [Paragraph(f"<b>{pesaje.get('nro_doc', '')}</b>",
                   s(28, bold=True, align=TA_CENTER, color=BLANCO))],
    ], colWidths=[W * 0.20])
    nro_cell.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), AZ),
        ("TOPPADDING",    (0,0),(-1,-1), 4),
        ("BOTTOMPADDING", (0,0),(-1,-1), 4),
        ("LEFTPADDING",   (0,0),(-1,-1), 4),
        ("RIGHTPADDING",  (0,0),(-1,-1), 4),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
    ]))

    # Ensamblar encabezado — las 3 columnas con mismo alto forzado
    cab = Table([[logo_cell, datos_cell, nro_cell]],
                colWidths=[W * 0.33, W * 0.47, W * 0.20],
                rowHeights=[HEADER_H])
    cab.setStyle(TableStyle([
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("LEFTPADDING",   (0,0),(-1,-1), 0),
        ("RIGHTPADDING",  (0,0),(-1,-1), 0),
        ("TOPPADDING",    (0,0),(-1,-1), 0),
        ("BOTTOMPADDING", (0,0),(-1,-1), 0),
        ("BACKGROUND",    (0,0),(0,0),   BLANCO),
        ("BACKGROUND",    (1,0),(1,0),   AZ_OSC),
        ("BACKGROUND",    (2,0),(2,0),   AZ),
    ]))
    e += [cab, Spacer(1, 1*mm)]

    # Sub-banda certificado
    cert     = certificado or {}
    nro_cert = cert.get("nro_certificado", "")
    vto_cert = cert.get("vencimiento", "")
    cert_str = ""
    if nro_cert:
        cert_str = f"Cert. Báscula: <b>{nro_cert}</b>"
    if vto_cert:
        cert_str += f"   |   Vencimiento: <b>{vto_cert[:10]}</b>"
    if cert_str:
        sub = Table([[Paragraph(cert_str, s(8, align=TA_CENTER, color=GRIS))]],
                    colWidths=[W])
        sub.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,-1), GRIS_L),
            ("TOPPADDING",    (0,0),(-1,-1), 3),
            ("BOTTOMPADDING", (0,0),(-1,-1), 3),
        ]))
        e += [sub]
    e += [Spacer(1, 4*mm)]

    # ════════════════════════════════════════════════════════════════════
    # BLOQUE DE PESOS
    # ════════════════════════════════════════════════════════════════════
    tara  = pesaje.get("peso_tara",  0) or 0
    bruto = pesaje.get("peso_bruto", 0) or 0
    neto  = pesaje.get("peso_neto",  0) or 0
    f_ing = (pesaje.get("fecha_ingreso") or "")[:16]
    f_eg  = (pesaje.get("fecha_egreso")  or "")[:16]

    if tipo == "egreso":
        # Tres columnas: Tara | Bruto | Neto
        wp = W * 0.57
        pesos = Table([
            [Paragraph("TARA",  s(8, bold=True, align=TA_CENTER, color=GRIS)),
             Paragraph("BRUTO", s(8, bold=True, align=TA_CENTER, color=AZ)),
             Paragraph("NETO",  s(8, bold=True, align=TA_CENTER, color=VERDE))],
            [Paragraph(f"<b>{_fmt(tara)}</b>",
                       s(13, bold=True, align=TA_CENTER)),
             Paragraph(f"<b>{_fmt(bruto)}</b>",
                       s(13, bold=True, align=TA_CENTER, color=AZ)),
             Paragraph(f"<b>{_fmt(neto)}</b>",
                       s(19, bold=True, align=TA_CENTER, color=VERDE))],
            [Paragraph("kg", s(8, align=TA_CENTER, color=GRIS)),
             Paragraph("kg", s(8, align=TA_CENTER, color=GRIS)),
             Paragraph("kg", s(8, bold=True, align=TA_CENTER, color=VERDE))],
        ], colWidths=[wp/3, wp/3, wp/3])
        pesos.setStyle(TableStyle([
            ("BOX",           (0,0),(-1,-1), 1, GRIS_L),
            ("LINEAFTER",     (0,0),(1,-1),  1, GRIS_L),
            ("BACKGROUND",    (0,0),(0,-1),  GRIS_F),
            ("BACKGROUND",    (1,0),(1,-1),  AZ_BG),
            ("BACKGROUND",    (2,0),(2,-1),  VERDE_BG),
            ("TOPPADDING",    (0,0),(-1,-1), 5),
            ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ]))

        wf = W - wp - 3
        fechas = Table([
            [Paragraph("FECHA INGRESO", s(8, bold=True, color=GRIS, align=TA_CENTER)),
             Paragraph("FECHA EGRESO",  s(8, bold=True, color=GRIS, align=TA_CENTER))],
            [Paragraph(f"<b>{f_ing}</b>", s(9, bold=True, align=TA_CENTER)),
             Paragraph(f"<b>{f_eg}</b>",  s(9, bold=True, align=TA_CENTER))],
        ], colWidths=[wf/2, wf/2])
        fechas.setStyle(TableStyle([
            ("BOX",           (0,0),(-1,-1), 1, GRIS_L),
            ("LINEAFTER",     (0,0),(0,-1),  1, GRIS_L),
            ("LINEBELOW",     (0,0),(-1,0),  0.5, GRIS_L),
            ("TOPPADDING",    (0,0),(-1,-1), 5),
            ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ]))

        combo = Table([[pesos, fechas]], colWidths=[wp + 2, wf + 1])
        combo.setStyle(TableStyle([
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
            ("LEFTPADDING",   (0,0),(-1,-1), 0),
            ("RIGHTPADDING",  (0,0),(-1,-1), 0),
        ]))
        e += [combo, Spacer(1, 4*mm)]

    else:
        # Ingreso: solo tara + fecha
        it = Table([
            [Paragraph("PESO TARA", s(8, bold=True, align=TA_CENTER, color=GRIS)),
             Paragraph("FECHA INGRESO", s(8, bold=True, align=TA_CENTER, color=GRIS))],
            [Paragraph(f"<b>{_fmt(tara)}</b>",
                       s(15, bold=True, align=TA_CENTER, color=AZ)),
             Paragraph(f"<b>{f_ing}</b>", s(10, bold=True, align=TA_CENTER))],
            [Paragraph("kg", s(8, align=TA_CENTER, color=GRIS)), Paragraph("", s(8))],
        ], colWidths=[W*0.35, W*0.65])
        it.setStyle(TableStyle([
            ("BOX",           (0,0),(-1,-1), 1, GRIS_L),
            ("LINEAFTER",     (0,0),(0,-1),  1, GRIS_L),
            ("LINEBELOW",     (0,0),(-1,0),  0.5, GRIS_L),
            ("BACKGROUND",    (0,0),(0,-1),  AZ_BG),
            ("TOPPADDING",    (0,0),(-1,-1), 5),
            ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ]))
        e += [it, Spacer(1, 4*mm)]

    # ════════════════════════════════════════════════════════════════════
    # SECCIONES DE DATOS
    # ════════════════════════════════════════════════════════════════════
    cw = [W*0.20, W*0.30, W*0.20, W*0.30]

    def sec(titulo, filas):
        hdr = Table([[Paragraph(titulo, s(9, bold=True, color=BLANCO))]],
                    colWidths=[W])
        hdr.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,-1), AZ),
            ("TOPPADDING",    (0,0),(-1,-1), 3),
            ("BOTTOMPADDING", (0,0),(-1,-1), 3),
            ("LEFTPADDING",   (0,0),(-1,-1), 8),
        ]))
        bod = Table(filas, colWidths=cw)
        bod.setStyle(TableStyle([
            ("LINEBELOW",     (0,0),(-1,-2), 0.3, GRIS_L),
            ("TOPPADDING",    (0,0),(-1,-1), 4),
            ("BOTTOMPADDING", (0,0),(-1,-1), 4),
            ("LEFTPADDING",   (0,0),(-1,-1), 8),
            ("BACKGROUND",    (0,0),(-1,-1), GRIS_F),
            ("BOX",           (0,0),(-1,-1), 0.5, GRIS_L),
        ]))
        return hdr, bod

    def r(l1, v1, l2="", v2=""):
        return [
            Paragraph(l1, s(8, color=GRIS)),
            Paragraph(f"<b>{v1}</b>" if v1 else "", s(9)),
            Paragraph(l2, s(8, color=GRIS)),
            Paragraph(f"<b>{v2}</b>" if v2 else "", s(9)),
        ]

    # Exportador / Transportista
    exp_nom  = pesaje.get("exportador",         "")
    exp_cuit = pesaje.get("cuit_exportador",    "")
    tr_nom   = pesaje.get("transportista",      "")
    tr_cuit  = pesaje.get("cuit_transportista", "")
    for blq in sec("EXPORTADOR / TRANSPORTISTA", [
        r("Exportador",    f"{exp_nom}  (CUIT {exp_cuit})",
          "Transportista", tr_nom),
    ]):
        e.append(blq)
    e.append(Spacer(1, 2*mm))

    # Vehículo / Chofer
    pat_cam = pesaje.get("patente_camion",    "")
    des_cam = pesaje.get("desc_camion",       "")
    pat_ac  = pesaje.get("patente_acoplado",  "")
    chofer  = pesaje.get("chofer",            "")
    tdoc    = pesaje.get("tipo_doc_chofer",   "DNI")
    ndoc    = pesaje.get("nro_doc_chofer",    "")
    nacion  = pesaje.get("nacionalidad_chofer", "")
    for blq in sec("VEHÍCULO / CHOFER", [
        r("Camión",   pat_cam,  "Marca / Modelo", des_cam),
        r("Acoplado", pat_ac,   "Chofer",         chofer),
        r("Tipo doc", tdoc,     "Nro / Nac.",     f"{ndoc} / {nacion}"),
    ]):
        e.append(blq)
    e.append(Spacer(1, 2*mm))

    # Carga / Identificación
    prod = pesaje.get("producto",            "")
    cont = pesaje.get("id_contenedor",       "")
    prec = pesaje.get("precintos",           "")
    dest = pesaje.get("destinacion_aduanera","")
    obs  = pesaje.get("observaciones",       "")

    filas_carga = [
        r("Contenedor", cont, "Producto", prod),
    ]
    if tipo == "egreso":
        filas_carga.append(
            r("Precintos", prec, "Dest. aduanera", dest)
        )
        if obs:
            filas_carga.append(
                r("Observaciones", obs, "", "")
            )
    for blq in sec("CARGA / IDENTIFICACIÓN", filas_carga):
        e.append(blq)

    e.append(Spacer(1, 6*mm))

    # ════════════════════════════════════════════════════════════════════
    # PIE
    # ════════════════════════════════════════════════════════════════════
    usr  = pesaje.get("usuario_egreso" if tipo == "egreso"
                      else "usuario_ingreso", "")
    pie = Table([[
        Paragraph("Sistema de Balanza v2.0 — ARGENSUN S.A.", s(7, color=GRIS)),
        Paragraph(f"Operador: <b>{usr}</b>", s(8, bold=True, align=TA_RIGHT)),
    ]], colWidths=[W * 0.6, W * 0.4])
    pie.setStyle(TableStyle([
        ("LINEABOVE",     (0,0),(-1,0), 0.5, GRIS_L),
        ("TOPPADDING",    (0,0),(-1,-1), 4),
        ("BOTTOMPADDING", (0,0),(-1,-1), 2),
    ]))
    e.append(pie)

    doc.build(e)
    return ruta


# ── Excel ─────────────────────────────────────────────────────────────────────

def exportar_excel(pesajes: list, ruta: str) -> str:
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise RuntimeError("Instala openpyxl: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pesajes"
    azul  = "1A5FA8"
    azul2 = "2C5F8A"
    gris  = "F5F5F5"
    borde = Border(
        left=Side("thin", color="CCCCCC"),  right=Side("thin", color="CCCCCC"),
        top= Side("thin", color="CCCCCC"),  bottom=Side("thin", color="CCCCCC"),
    )

    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value     = "REGISTRO DE PESAJES"
    c.font      = Font(bold=True, color="FFFFFF", size=13)
    c.fill      = PatternFill("solid", fgColor=azul)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:N2")
    ws["A2"].value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A2"].font  = Font(italic=True, color="888888", size=9)
    ws["A2"].alignment = Alignment(horizontal="center")

    encab = [
        ("Doc",10), ("Tipo",10), ("F.Egreso",18), ("Exportador",22),
        ("Transportista",22), ("Patente",10), ("Camion",14), ("Acoplado",10),
        ("Chofer",20), ("Producto",14), ("Tara",12), ("Bruto",12),
        ("Neto (kg)",12), ("Dest.",8),
    ]
    for i, (t, w) in enumerate(encab, 1):
        c2 = ws.cell(row=3, column=i, value=t)
        c2.font      = Font(bold=True, color="FFFFFF", size=9)
        c2.fill      = PatternFill("solid", fgColor=azul2)
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border    = borde
        ws.column_dimensions[c2.column_letter].width = w
    ws.row_dimensions[3].height = 18

    for ri, p in enumerate(pesajes, 4):
        fondo = PatternFill("solid", fgColor=gris if ri % 2 == 0 else "FFFFFF")
        vals = [
            p.get("nro_doc", ""),         p.get("tipo_pesaje", ""),
            (p.get("fecha_egreso") or "")[:16], p.get("exportador", ""),
            p.get("transportista", ""),   p.get("patente_camion", ""),
            p.get("desc_camion", ""),     p.get("patente_acoplado", ""),
            p.get("chofer", ""),          p.get("producto", ""),
            p.get("peso_tara",  0),       p.get("peso_bruto", 0),
            p.get("peso_neto",  0),       p.get("destinacion_aduanera", ""),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.fill   = fondo
            cell.border = borde
            cell.font   = Font(size=9)
            if ci in (11, 12, 13):
                cell.number_format = "#,##0.0"
                cell.alignment     = Alignment(horizontal="right")

    tr = len(pesajes) + 4
    ws.cell(row=tr, column=12, value="TOTAL NETO:").font = Font(bold=True, size=9)
    c3 = ws.cell(row=tr, column=13,
                 value=sum(p.get("peso_neto", 0) for p in pesajes))
    c3.font           = Font(bold=True, size=9)
    c3.number_format  = "#,##0.0"
    c3.alignment      = Alignment(horizontal="right")

    ws.freeze_panes = "A4"
    os.makedirs(os.path.dirname(os.path.abspath(ruta)), exist_ok=True)
    wb.save(ruta)
    return ruta
